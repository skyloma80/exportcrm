"""Parser for structured data from Excel files (.xlsx, .xls).

Extracts document type, customer info, line items, totals, and
payment/bank details from Proforma Invoices, Purchase Orders,
and Quotations.
"""
from __future__ import annotations
from typing import Any


def parse_excel(file_path: str) -> dict:
    """Parse an Excel file and extract business data.

    Args:
        file_path: Path to the .xlsx or .xls file

    Returns:
        Dict with parsed data in common format:
        {
            "doc_type": "pi" | "po" | "quotation" | "unknown",
            "doc_code": str,
            "date": str,
            "customer": {"name", "address", "tax_id", "contact"},
            "supplier": {"name", "address"},  # only for PO
            "items": [{"product_name", "part_number", "quantity",
                       "unit", "unit_price", "amount", "description"}],
            "currency": str,
            "total_amount": float,
            "payment_terms": str,
            "bank_info": str,
            "incoterm": str,
            "port_of_loading": str,
            "port_of_destination": str,
        }
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=False))

    text_rows = []
    for row in rows:
        line = " ".join(str(c.value or "") for c in row).strip()
        text_rows.append(line)

    doc_type = _detect_doc_type(text_rows)
    doc_code = _extract_doc_code(text_rows, doc_type)
    date = _extract_date(text_rows)
    customer = _extract_customer(text_rows)
    supplier = _extract_supplier(text_rows) if doc_type == "po" else {}
    items = _extract_items(ws)
    currency = _extract_currency(text_rows)
    total_amount = _extract_total(text_rows)
    payment_terms = _extract_field(text_rows, ["payment terms", "付款条件",
                                                "payment", "terms of payment"])
    bank_info = _extract_bank_info(text_rows) if doc_type == "pi" else ""
    incoterm = _extract_field(text_rows, ["incoterm", "trade term",
                                           "delivery term", "贸易条款"])
    port_of_loading = _extract_field(text_rows, ["port of loading",
                                                  "loading port", "装货港"])
    port_of_destination = _extract_field(text_rows, ["port of destination",
                                                      "destination port", "目的港"])

    wb.close()

    return {
        "doc_type": doc_type,
        "doc_code": doc_code,
        "date": date,
        "customer": customer,
        "supplier": supplier,
        "items": items,
        "currency": currency,
        "total_amount": total_amount,
        "payment_terms": payment_terms,
        "bank_info": bank_info,
        "incoterm": incoterm,
        "port_of_loading": port_of_loading,
        "port_of_destination": port_of_destination,
    }


def _detect_doc_type(text_rows: list[str]) -> str:
    """Detect document type from keywords."""
    combined = " ".join(text_rows).lower()
    if any(k in combined for k in ["proforma invoice", "pi-", "形式发票"]):
        return "pi"
    if any(k in combined for k in ["purchase order", "po-", "采购订单"]):
        return "po"
    if any(k in combined for k in ["quotation", "quote", "报价单"]):
        return "quotation"
    if any(k in combined for k in ["commercial invoice", "invoice", "发票"]):
        return "invoice"
    return "unknown"


def _extract_doc_code(text_rows: list[str], doc_type: str) -> str:
    """Extract document number."""
    combined = "\n".join(text_rows[:20])
    import re

    patterns = {
        "pi": r"PI[-:=\s]*([A-Z0-9\-/]+)",
        "po": r"PO[-:=\s]*([A-Z0-9\-/]+)",
        "quotation": r"(?:QUO|QTN|QT)[-:=\s]*([A-Z0-9\-/]+)",
        "invoice": r"(?:INV|发票)[-:=\s]*([A-Z0-9\-/]+)",
    }

    pattern = patterns.get(doc_type, r"(?:No|编号|CODE)[\.:：\s]*([A-Z0-9\-/]+)")
    m = re.search(pattern, combined, re.IGNORECASE)
    return m.group(1).strip() if m else ""


def _extract_date(text_rows: list[str]) -> str:
    """Extract document date."""
    combined = "\n".join(text_rows[:20])
    import re

    m = re.search(r"(?:date|日期|issu(?:e|ed))[\s\.:：]*(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})",
                  combined, re.IGNORECASE)
    if m:
        return m.group(1)

    m = re.search(r"(\d{2}[-/.]\d{2}[-/.]\d{4})", combined)
    if m:
        return m.group(1)

    return ""


def _extract_customer(text_rows: list[str]) -> dict:
    """Extract customer information from document."""
    combined = "\n".join(text_rows[:30])
    lines = combined.split("\n")

    customer = {"name": "", "address": "", "tax_id": "", "contact": ""}
    found_section = False

    key_indicators = ["bill to", "ship to", "customer", "buyer",
                      " sold to", "consignee", "客户", "买方"]

    for i, line in enumerate(lines):
        lower = line.lower().strip()
        if any(k in lower for k in key_indicators):
            found_section = True
            # The next non-empty line after the label is likely the name
            for j in range(i + 1, min(i + 5, len(lines))):
                if lines[j].strip() and ":" not in lines[j]:
                    customer["name"] = lines[j].strip()
                    break
            continue

        if found_section and line.strip():
            # Collect address lines until we hit another section
            if customer["name"] and line.strip() != customer["name"]:
                if not customer["address"]:
                    customer["address"] = line.strip()
                else:
                    customer["address"] += "\n" + line.strip()

    # Extract tax ID / VAT
    import re
    m = re.search(r"(?:tax|vat|tin)[\s\.:：]*([A-Z0-9\-]+)",
                  combined, re.IGNORECASE)
    if m:
        customer["tax_id"] = m.group(1)

    # Clean up
    for key in customer:
        if isinstance(customer[key], str):
            customer[key] = customer[key].strip()

    return customer


def _extract_supplier(text_rows: list[str]) -> dict:
    """Extract supplier info (for PO documents)."""
    combined = "\n".join(text_rows[:30])
    lines = combined.split("\n")

    supplier = {"name": "", "address": ""}
    found_section = False

    key_indicators = ["supplier", "vendor", "seller", "卖方", "供应商"]

    for i, line in enumerate(lines):
        lower = line.lower().strip()
        if any(k in lower for k in key_indicators):
            found_section = True
            for j in range(i + 1, min(i + 5, len(lines))):
                if lines[j].strip() and ":" not in lines[j]:
                    supplier["name"] = lines[j].strip()
                    break
            continue

        if found_section and line.strip():
            if supplier["name"] and line.strip() != supplier["name"]:
                if not supplier["address"]:
                    supplier["address"] = line.strip()
                else:
                    supplier["address"] += "\n" + line.strip()

    for key in supplier:
        if isinstance(supplier[key], str):
            supplier[key] = supplier[key].strip()

    return supplier


def _extract_items(ws) -> list[dict]:
    """Extract line items from worksheet.

    Scans rows for table headers (quantity, unit price, amount, etc.)
    and extracts data rows below.
    """
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    header_row_idx = _find_header_row(rows)
    if header_row_idx < 0:
        return []

    items = []
    for row in rows[header_row_idx + 1:]:
        vals = [str(v or "").strip() for v in row]

        if all(not v for v in vals):
            break

        item = _parse_item_row(vals)
        if item and item.get("quantity", 0) > 0:
            items.append(item)

    return items


def _find_header_row(rows: list[tuple]) -> int:
    """Find the row containing table headers."""
    header_keywords = ["quantity", "qty", "unit price", "amount",
                       "description", "product", "part number",
                       "数量", "单价", "金额", "产品"]

    for i, row in enumerate(rows):
        combined = " ".join(str(v or "").lower() for v in row)
        match_count = sum(1 for k in header_keywords if k in combined)
        if match_count >= 3:
            return i
    return -1


def _parse_item_row(vals: list[str]) -> dict | None:
    """Parse a single item row from table."""
    import re

    item = {
        "product_name": "",
        "part_number": "",
        "description": "",
        "quantity": 0,
        "unit": "",
        "unit_price": 0.0,
        "amount": 0.0,
    }

    # Try to find numeric values in the row
    nums = []
    for v in vals:
        clean = v.replace(",", "").replace("$", "").replace("€", "")
        try:
            num = float(clean)
            nums.append(num)
        except ValueError:
            pass

    if len(nums) >= 3:
        item["quantity"] = int(nums[0]) if nums[0] == int(nums[0]) else nums[0]
        item["unit_price"] = nums[-2]
        item["amount"] = nums[-1]
    elif len(nums) == 2:
        item["quantity"] = int(nums[0]) if nums[0] == int(nums[0]) else nums[0]
        item["unit_price"] = nums[1]
        item["amount"] = nums[0] * nums[1]
    elif len(nums) == 1:
        item["quantity"] = int(nums[0]) if nums[0] == int(nums[0]) else nums[0]

    # First non-numeric value is likely the product description
    text_vals = [v for v in vals if v and not _is_numeric(v)]
    if text_vals:
        item["description"] = text_vals[0]
        item["product_name"] = text_vals[0]

    return item


def _is_numeric(v: str) -> bool:
    try:
        float(v.replace(",", "").replace("$", "").replace("€", ""))
        return True
    except ValueError:
        return False


def _extract_currency(text_rows: list[str]) -> str:
    """Extract currency code."""
    combined = "\n".join(text_rows[:40])
    import re

    currencies = ["USD", "EUR", "CNY", "GBP", "JPY", "HKD", "AUD", "CAD", "SGD"]
    for cur in currencies:
        if re.search(rf"\b{cur}\b", combined, re.IGNORECASE):
            return cur
    return "USD"


def _extract_total(text_rows: list[str]) -> float:
    """Extract total amount from document."""
    combined = "\n".join(text_rows)
    import re

    patterns = [
        r"(?:total|合计|总计)[\s\.:：]*[$€£]*\s*([\d,]+\.?\d*)",
        r"(?:amount due|grand total)[\s\.:：]*[$€£]*\s*([\d,]+\.?\d*)",
        r"[$€£]*\s*([\d,]+\.\d{2})\s*$",
    ]

    for pattern in patterns:
        matches = re.findall(pattern, combined, re.IGNORECASE | re.MULTILINE)
        if matches:
            try:
                return float(matches[-1].replace(",", ""))
            except ValueError:
                pass
    return 0.0


def _extract_field(text_rows: list[str], keywords: list[str]) -> str:
    """Extract a field value following a keyword."""
    combined = "\n".join(text_rows[:50])
    lines = combined.split("\n")
    import re

    for i, line in enumerate(lines):
        lower = line.lower()
        for kw in keywords:
            if kw in lower:
                parts = re.split(r"[:：\s]+", line, maxsplit=1)
                if len(parts) > 1:
                    return parts[1].strip()
                if i + 1 < len(lines):
                    return lines[i + 1].strip()
    return ""


def _extract_bank_info(text_rows: list[str]) -> str:
    """Extract bank/remittance information."""
    combined = "\n".join(text_rows)
    lines = combined.split("\n")
    import re

    bank_lines = []
    in_bank_section = False
    bank_keywords = ["bank", "remittance", "account", "iban", "swift",
                     "银行", "汇款", "账户"]

    for line in lines:
        lower = line.lower()
        if any(k in lower for k in bank_keywords):
            in_bank_section = True

        if in_bank_section:
            bank_lines.append(line.strip())
            if len(bank_lines) > 15:
                break

    return "\n".join(bank_lines).strip()
